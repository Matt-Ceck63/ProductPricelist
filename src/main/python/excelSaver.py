from openpyxl import *
import os.path

# TODO: redefine as a function in main class

class ExcelSaver:

    def __init__(self, path):

        # TODO: Make context manager
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sheet"

    def save(self, data):

        for i, h in enumerate(data.columns, start=1):
            self.ws.cell(row=1, column=i,).value = h

        for r, row_values in enumerate(data.iterrows(), start=2):
            for c, cell in enumerate(row_values[1], start=1):

                self.ws.cell(row=r, column=c).value = cell

        for i, h in enumerate(data.columns):
            column_width = 15 # max([len(str(val)) for val in list(data[h])] + [len(str(h))]) + 5
            self.ws.column_dimensions[self.get_column_letter(i)].width = column_width

        for i, row in enumerate(self.ws.iter_rows()):
            if i == 0:
                for cell in row:
                    cell.alignment = cell.alignment.copy(wrapText=True, horizontal='center', vertical='center')

        # TODO: save to desktop in windows
        self.wb.save(filename="TESTEXCEL.xlsx")

    def get_column_letter(self, index):
        columns = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + ["AA", "AB", "AC", "AD", "AE"]
        return columns[index]
